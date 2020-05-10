from tkinter import *
from PIL import ImageTk, Image
import openpyxl
#import os
#os.getcwd()

wb_selections = []
sh_selections = []
duties = []
options = ['Select One']

#Key button functions
def click_1():
    text_1 = textentry_1.get()+".xlsx"
    output.delete(0.0, END)
    try:
        a_wb = openpyxl.load_workbook(text_1)
        output.insert(END, "Your available sheets within this workbook are {}".format(a_wb.sheetnames))
        wb_selections.append(a_wb)
    except:
        wb_error = "Please enter a valid workbook name.\n.\n.\n.\n. Names ARE case sensitive."
        output.insert(END, wb_error)
    else:
        wb_selections.append(a_wb)
        Label(window, text="-----Enter the name of an available sheet from the list below-----", bg="black", fg="white", font="none 12 bold").grid(row=1, column=0, sticky=W)

def click_2():
    text_2 = textentry_2.get()
    output.delete(0.0, END)
    wb = wb_selections[0]
    try:
        sh = wb[text_2]
        output.insert(END, "You selected the sheet named '{}'.\n.\n.\n.\n.".format(text_2))
    except:
        sh_error = "Please enter a valid sheet name. Your available names are: {}\n.\n.\n.\n Names ARE case sensitive".format(wb_selections[-1].sheetnames)
        output.insert(END, sh_error)
    else:
        sh_selections.append(sh)
        duty_template()
        feat_pers_list()
        #a_dut = duties[-1]
        #output.insert(END, "Your sheet has a duties roster of the following: \n.\n {}".format(a_dut))

def click_3():

    try:
        output.delete(0.0, END)
        date = int(textentry_3.get())
        if date <= 31:
            output.insert(END, "{} worked on Day: {}".format(duties[0][date], date))
    except:
        output.insert(END, "Please enter a valid day between 1-31")

#Main Window
window = Tk()
window.title("QDuties")
window.geometry("800x600")
window.resizable(0,0)

#window.configure(background="black")
background_image = PhotoImage(file="Ma.gif")
background_label = Label(window, image=background_image)
background_label.place(x=0, y=0, relwidth=1, relheight=1)

#Function for displayiong the names featured on the duties roster
def feat_pers_list():
    feat_pers = []

    days = duties[0].items()
    for day in days:
        if day[1][1] not in feat_pers and day[1][1] != None :
            feat_pers.append(day[1][1])
            options.append(day[1][1])
        elif day[1][0] not in feat_pers and day[1][0] != None:
            feat_pers.append(day[1][0])
            options.append(day[1][0])
            u_feat_pers = set(feat_pers)
    output.insert(END, "\nThe persons on this duties are {}\n.\n.\n.\n {}.".format(u_feat_pers,options))
    dropdown = OptionMenu(window, clicked, *options)
    dropdown.grid(row=5, column=0, sticky=W)

#Function for showing the persons duty on specific days
def show():
    output.delete(0.0, END)
    pers = clicked.get()
    duty_days = []
    days = duties[0].items()
    for day in days:
        if day[1][1] == pers or day[1][0] == pers:
            duty_days.append(day[0])
    output.insert(END, "\n.\n.\n.{} duty days are as follows: {}".format(pers, duty_days))


    #Label(window, text=clicked.get(), bg="black", fg="white", font="none 12 bold").grid(row=4, column=0, sticky=W)


#Button to show results for dropdown selection
clicked = StringVar()
clicked.set(options[0])
Button(window, text="Check Person", width=14, command=show).grid(row=5, column=2, sticky=None)
#output.insert(END, "\nThis is a test {}\n.".format(pers))

#Label for user to select workbook
Label (window, text="---Enter the name of the excel workbook with a '.xlsx' extension---", bg="black", fg="white", font="none 12 bold") .grid(row=0, column=0, sticky=NW)
#Text entry box for workbook selection
textentry_1 = Entry(window, width=30, bg="white")
textentry_1.grid(row=0, column=2, sticky=None)
#submit button for workbook selection
Button (window, text="Check Workbook", width=14, command=click_1) .grid(row=0, column=5, sticky=None)

#Output Window
output = Text(window, width=60, height=20, wrap=WORD, background="white")
output.grid(row=3, column=0, sticky=None)
Label(window, text="QDuty Results Window", bg="black", fg="white", font="none 12 bold").grid(row=3, column=2, sticky=E)

#Textentry box For user to choose a sheet
textentry_2 = Entry(window, width=30, bg="white")
textentry_2.grid(row=1, column=2, sticky=None)
Button(window, text="Check Sheet", width=14, command=click_2).grid(row=1, column=5, sticky=None)

#Textentry box For user to imput a date/dates and find out who was duty
textentry_3 = Entry(window, width=10, bg="white")
textentry_3.grid(row=8, column=2)
Button(window, text="Check Dates", width=14, command=click_3).grid(row=9, column=2)

#Function to create duties template
def duty_template():
    sh = sh_selections[0]
    template = {sh['B5'].value:[sh['C5'].value,sh['C6'].value],sh['D5'].value:[sh['E5'].value,sh['E6'].value],
                sh['F5'].value:[sh['G5'].value,sh['G6'].value],sh['H5'].value:[sh['I5'].value,sh['I6'].value],
                sh['J5'].value:[sh['K5'].value,sh['K6'].value],sh['L5'].value:[sh['M5'].value,sh['M6'].value],
                sh['N5'].value:[sh['O5'].value,sh['O6'].value],sh['B7'].value:[sh['C7'].value,sh['C8'].value],
                sh['D7'].value:[sh['E7'].value,sh['E8'].value],sh['F7'].value:[sh['G7'].value,sh['G8'].value],
                sh['H7'].value:[sh['I7'].value,sh['I8'].value],sh['J7'].value:[sh['K7'].value,sh['K8'].value],
                sh['L7'].value:[sh['M7'].value,sh['M8'].value],sh['N7'].value:[sh['O7'].value,sh['O8'].value],
                sh['B9'].value:[sh['C9'].value,sh['C10'].value],sh['D9'].value:[sh['E9'].value,sh['E10'].value],
                sh['F9'].value:[sh['G9'].value,sh['G10'].value],sh['H9'].value:[sh['I9'].value,sh['I10'].value],
                sh['J9'].value:[sh['K9'].value,sh['K10'].value],sh['L9'].value:[sh['M9'].value,sh['M10'].value],
                sh['N9'].value:[sh['O9'].value,sh['O10'].value],sh['B11'].value:[sh['C11'].value,sh['C12'].value],
                sh['D11'].value:[sh['E11'].value,sh['E12'].value],sh['F11'].value:[sh['G11'].value,sh['G12'].value],
                sh['H11'].value:[sh['I11'].value,sh['I12'].value],sh['J11'].value:[sh['K11'].value,sh['K12'].value],
                sh['L11'].value:[sh['M11'].value,sh['M12'].value],sh['N11'].value:[sh['O11'].value,sh['O12'].value],
                sh['B13'].value:[sh['C13'].value,sh['C14'].value],sh['D13'].value:[sh['E13'].value,sh['E14'].value],
                sh['F13'].value:[sh['G13'].value,sh['G14'].value],sh['H13'].value:[sh['I13'].value,sh['I14'].value],
                sh['J13'].value:[sh['K13'].value,sh['K14'].value],sh['L13'].value:[sh['M13'].value,sh['M14'].value],
                sh['N13'].value:[sh['O13'].value,sh['O14'].value],sh['B15'].value:[sh['C15'].value,sh['C16'].value],
                sh['D15'].value:[sh['E15'].value,sh['E16'].value],sh['F15'].value:[sh['G15'].value,sh['G16'].value],
                sh['H15'].value:[sh['I15'].value,sh['I16'].value],sh['J15'].value:[sh['K15'].value,sh['K16'].value],
                sh['L15'].value:[sh['M15'].value,sh['M16'].value],sh['N15'].value:[sh['O15'].value,sh['O16'].value]}
    dut = {k: v for k, v in template.items() if k is not None}
    duties.append(dut)

window.mainloop()

#09.05.2020








